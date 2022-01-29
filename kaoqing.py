# from email.mime import image
from math import ceil
from openpyxl import load_workbook
from chinese_calendar import holidays, is_workday, is_holiday, is_in_lieu
import calendar
import os
import pandas as pd
from employee import Employee


def get_monthrange(year,month):
    """获取指定月份所有日期的一个生成器

    Args:
        year (int)): 指定年份
        month (int): 指定月份

    Returns:
        generator: 所有日期的一个生成器
    """
    _c = calendar.Calendar()
    _month = _c.itermonthdates(year,month)
    monthrange = []
    [monthrange.append(each) for each in _month if each.month==month]
    return monthrange

def get_workdays(monthrange):
    """计算并返回指定月份的工作日数量

    Args:
        monthrange (generator): 月份生成器

    Returns:
        int: 工作日数量
    """
    workday_count = 0
    for each in monthrange:
        if is_workday(each):
            workday_count +=1
    return workday_count

def get_holidays_count(monthrange):
    """计算并返回指定月份的休息日和节假日总和

    Args:
        monthrange (generator): 指定月份的生成器

    Returns:
        int: 总节假日和休息日的数量
    """
    holiday_count = 0
    for each in monthrange:
        if is_holiday(each):
            holiday_count += 1
    return holiday_count

def get_six_workdays(monthrange):
    _count = 0
    for each in monthrange:
        if is_workday(each) or each.isoweekday() == 6:
            _count +=1
    return _count

def get_six_holidays(monthrange):
    holiday_count = 0
    for each in monthrange:
        if is_holiday(each) and each.isoweekday() != 6:
            holiday_count += 1
    return holiday_count

def get_legal_holidays_count(monthrange):
    _count = 0
    for each in monthrange:
        if is_holiday(each) and is_in_lieu(each):
            _count += 1
    return _count

def get_workday_info(year,month):
    _info = {}
    monthrange = get_monthrange(2021,12)
    _info['workday_count'] = get_workdays(monthrange)
    _info['holiday_count'] = get_holidays_count(monthrange)
    _info['legal_holiday'] = get_legal_holidays_count(monthrange)
    _info['six_workdays'] = get_six_workdays(monthrange)
    _info['six_holidays'] = get_six_holidays(monthrange)
    return _info

def update_basic_info(ws, year, month, info):
    ws['b1'] = month
    ws['b2'] = month if month<12 else 1
    ws['b3'] = year
    ws['b5'] = info['workday_count']
    ws['b6'] = info['holiday_count']
    ws['b7'] = info['legal_holiday']
    ws['b8'] = info['six_workdays']
    ws['b9'] = info['six_holidays']

def get_filename(name_prefix):
    _files = os.listdir('.//考勤统计表')
    filename = './/考勤统计表//'
    for each in _files:
        if each.startswith(name_prefix):
            filename += each
    return filename

def get_employees():
    filename = get_filename('广丰惠州_月度汇总')
    wb = load_workbook(filename)
    ws = wb['月度汇总']
    iter_employees = ws.iter_cols(max_col=1,min_row=5,max_row=ws.max_row,values_only=True)
    employees_list = {}
    for each in iter_employees:
        for i in each:
            if i != '陈江':
                employees_list[i] = Employee(i)
            if i == '刘光华' or i == '林锦萍' or i == '张辉':
                employees_list[i].set_six_workdays_mode()
    return employees_list

def get_kaoqing_details(filename):
    wb = load_workbook(filename)
    ws = wb.active
    title = []
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=37,values_only=True):
        for cell in row:
            title.append(cell)
    ws.delete_rows(1,4)
    ws.delete_cols(38,ws.max_column-37)
    return pd.DataFrame(ws.values, columns=title)

def update_employees_kaoqing_info(em, workday_info, dataframe):
    if em.get_workdays_mode() == True:
        em.set_yingchu_days(workday_info['six_workdays'])
    else:
        em.set_yingchu_days(workday_info['workday_count'])
    penson = dataframe.loc[dataframe['姓名']==em.get_name()]
    for row in penson.itertuples():
        if row[9] != '休息' and row[9] != '':
            if row[23] == '1':
                em.add_actual_workdays()
            elif row[37] is not None:
                if int(row[37]) >= 7:
                    em.add_actual_workdays()
    pass

def update_hengke_info(ws, year, month, employees_list):
    ws['a2'] = '惠州恒科房地产开发有限公司%s年%s月份员工考勤确认统计表' % (year,month)
    ws['p3'] = '制表日期:%s-%s-1' % (year,month)

def main():
    year = 2021
    month = 12
    wb = load_workbook('.//考勤统计表//考勤确认表.xlsx',data_only=True)

    # 更新基础信息表
    workday_info = get_workday_info(year,month)
    update_basic_info(wb['基础信息表'], year, month, workday_info)

    # 更新员工考勤信息
    employees_list = get_employees()
    df = get_kaoqing_details(get_filename('广丰惠州_每日统计'))
    for each in employees_list.values():
        update_employees_kaoqing_info(each, workday_info, df)

    # 更新恒科表单
    update_hengke_info(wb['恒科'], year, month, employees_list)


    # 保存
    wb.save('.//考勤统计表//%s月考勤确认表.xlsx' % month)

if __name__ == '__main__':
    main()
    